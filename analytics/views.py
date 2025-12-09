from django.shortcuts import render, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count
from django.http import HttpResponse
from forms_app.models import FeedbackForm, Question, Response, MCQOption
from core.models import School, Department, Course
from accounts.models import Student, StudentCourse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

@staff_member_required
def analytics_dashboard(request):
    schools = School.objects.all()
    selected_school = request.GET.get('school')
    selected_department = request.GET.get('department')
    selected_course = request.GET.get('course')
    
    departments = Department.objects.none()
    courses = Course.objects.none()
    forms = FeedbackForm.objects.none()
    
    if selected_school:
        departments = Department.objects.filter(school_id=selected_school)
    
    if selected_department:
        courses = Course.objects.filter(department_id=selected_department)
    
    if selected_course:
        forms = FeedbackForm.objects.filter(course_id=selected_course).select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        )
    
    context = {
        'schools': schools,
        'departments': departments,
        'courses': courses,
        'forms': forms,
        'selected_school': selected_school,
        'selected_department': selected_department,
        'selected_course': selected_course,
    }
    return render(request, 'analytics/dashboard.html', context)

@staff_member_required
def form_results(request, form_id):
    form = get_object_or_404(
        FeedbackForm.objects.select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        ), 
        id=form_id
    )
    
    questions = form.questions.all().prefetch_related('options')
    total_submissions = form.submissions.count()
    
    results = []
    
    for question in questions:
        question_data = {
            'question': question,
            'total_responses': 0,
            'data': []
        }
        
        if question.question_type == 'mcq':
            # Get all options for this question first
            all_options = question.options.all().order_by('order')
            
            # Get response counts for each option
            option_counts = Response.objects.filter(
                question=question,
                mcq_answer__isnull=False
            ).values(
                'mcq_answer__id',
                'mcq_answer__option_text',
                'mcq_answer__order'
            ).annotate(
                count=Count('id')
            ).order_by('mcq_answer__order')
            
            # Convert to dict for easier lookup
            counts_dict = {item['mcq_answer__id']: item for item in option_counts}
            
            # Build complete data including options with 0 responses
            complete_data = []
            for option in all_options:
                if option.id in counts_dict:
                    complete_data.append(counts_dict[option.id])
                else:
                    # Add option with 0 count
                    complete_data.append({
                        'mcq_answer__id': option.id,
                        'mcq_answer__option_text': option.option_text,
                        'mcq_answer__order': option.order,
                        'count': 0
                    })
            
            question_data['total_responses'] = sum(item['count'] for item in complete_data)
            question_data['data'] = complete_data
            
        elif question.question_type == 'text':
            # Get all text responses with student info
            text_responses = Response.objects.filter(
                question=question,
                text_answer__isnull=False
            ).exclude(
                text_answer=''
            ).select_related(
                'submission__student'
            ).values_list(
                'text_answer',
                'submission__student__name',
                'submission__submitted_at'
            ).order_by('-submission__submitted_at')
            
            question_data['total_responses'] = text_responses.count()
            question_data['data'] = list(text_responses)
        
        results.append(question_data)
    
    context = {
        'form': form,
        'results': results,
        'total_submissions': total_submissions,
    }
    return render(request, 'analytics/form_results.html', context)

@staff_member_required
def export_form_results(request, form_id):
    """Export form results to Excel file"""
    form = get_object_or_404(
        FeedbackForm.objects.select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        ), 
        id=form_id
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary information
    ws_summary['A1'] = "Feedback Report"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    ws_summary['A3'] = "Form Title:"
    ws_summary['B3'] = form.title
    ws_summary['A4'] = "Course:"
    ws_summary['B4'] = f"{form.course.code} - {form.course.name}"
    ws_summary['A5'] = "Teacher:"
    ws_summary['B5'] = form.teacher.name
    ws_summary['A6'] = "Department:"
    ws_summary['B6'] = form.course.department.name
    ws_summary['A7'] = "School:"
    ws_summary['B7'] = form.course.department.school.name
    ws_summary['A8'] = "Total Submissions:"
    ws_summary['B8'] = form.submissions.count()
    ws_summary['A9'] = "Generated On:"
    ws_summary['B9'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Make labels bold
    for row in range(3, 10):
        ws_summary[f'A{row}'].font = Font(bold=True)
    
    # MCQ Results Sheet
    ws_mcq = wb.create_sheet("MCQ Results")
    
    # Headers
    ws_mcq['A1'] = "Question"
    ws_mcq['B1'] = "Option"
    ws_mcq['C1'] = "Count"
    ws_mcq['D1'] = "Percentage"
    
    for col in ['A', 'B', 'C', 'D']:
        ws_mcq[f'{col}1'].fill = header_fill
        ws_mcq[f'{col}1'].font = header_font
        ws_mcq[f'{col}1'].border = border
        ws_mcq[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 2
    questions = form.questions.filter(question_type='mcq').prefetch_related('options')
    
    for question in questions:
        # Get response counts
        option_counts = Response.objects.filter(
            question=question,
            mcq_answer__isnull=False
        ).values(
            'mcq_answer__option_text'
        ).annotate(
            count=Count('id')
        ).order_by('mcq_answer__order')
        
        total_responses = sum(item['count'] for item in option_counts)
        
        # Write question and options
        first_row = current_row
        for idx, item in enumerate(option_counts):
            if idx == 0:
                ws_mcq[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
                ws_mcq[f'A{current_row}'].font = Font(bold=True)
            
            ws_mcq[f'B{current_row}'] = item['mcq_answer__option_text']
            ws_mcq[f'C{current_row}'] = item['count']
            
            if total_responses > 0:
                percentage = (item['count'] / total_responses) * 100
                ws_mcq[f'D{current_row}'] = f"{percentage:.1f}%"
            else:
                ws_mcq[f'D{current_row}'] = "0%"
            
            # Apply borders
            for col in ['A', 'B', 'C', 'D']:
                ws_mcq[f'{col}{current_row}'].border = border
            
            current_row += 1
        
        # Merge question cells
        if len(option_counts) > 1:
            ws_mcq.merge_cells(f'A{first_row}:A{current_row-1}')
        
        current_row += 1  # Empty row between questions
    
    # Text Responses Sheet
    ws_text = wb.create_sheet("Text Responses")
    
    # Headers
    ws_text['A1'] = "Question"
    ws_text['B1'] = "Student"
    ws_text['C1'] = "Response"
    ws_text['D1'] = "Submitted On"
    
    for col in ['A', 'B', 'C', 'D']:
        ws_text[f'{col}1'].fill = header_fill
        ws_text[f'{col}1'].font = header_font
        ws_text[f'{col}1'].border = border
        ws_text[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 2
    text_questions = form.questions.filter(question_type='text')
    
    for question in text_questions:
        responses = Response.objects.filter(
            question=question,
            text_answer__isnull=False
        ).exclude(
            text_answer=''
        ).select_related(
            'submission__student'
        ).order_by('submission__submitted_at')
        
        for response in responses:
            ws_text[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
            ws_text[f'B{current_row}'] = response.submission.student.name
            ws_text[f'C{current_row}'] = response.text_answer
            ws_text[f'D{current_row}'] = response.submission.submitted_at.strftime("%Y-%m-%d %H:%M")
            
            # Apply borders
            for col in ['A', 'B', 'C', 'D']:
                ws_text[f'{col}{current_row}'].border = border
            
            # Wrap text for response column
            ws_text[f'C{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            current_row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 50
    
    ws_mcq.column_dimensions['A'].width = 50
    ws_mcq.column_dimensions['B'].width = 40
    ws_mcq.column_dimensions['C'].width = 12
    ws_mcq.column_dimensions['D'].width = 12
    
    ws_text.column_dimensions['A'].width = 50
    ws_text.column_dimensions['B'].width = 25
    ws_text.column_dimensions['C'].width = 60
    ws_text.column_dimensions['D'].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Feedback_{form.course.code}_{form.teacher.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
def export_students_list(request):
    """Export all registered students to Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Students"
    
    # Styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['S.No', 'Roll Number', 'Name', 'School', 'Department', 'Enrolled Courses', 'Registration Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get all students with related data
    students = Student.objects.filter(is_staff=False).select_related(
        'school', 'department'
    ).prefetch_related('enrolled_courses__course').order_by('school', 'department', 'name')
    
    # Write data
    row_num = 2
    for idx, student in enumerate(students, 1):
        # Get enrolled courses
        courses = student.enrolled_courses.all()
        course_list = ', '.join([f"{sc.course.code}" for sc in courses]) if courses.exists() else 'None'
        
        # Write row
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=student.roll_number)
        ws.cell(row=row_num, column=3, value=student.name)
        ws.cell(row=row_num, column=4, value=student.school.name if student.school else 'N/A')
        ws.cell(row=row_num, column=5, value=student.department.name if student.department else 'N/A')
        ws.cell(row=row_num, column=6, value=course_list)
        ws.cell(row=row_num, column=7, value=student.date_joined.strftime("%Y-%m-%d"))
        
        # Apply borders
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 18
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Student Registration Summary"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    ws_summary['A3'] = "Total Students:"
    ws_summary['B3'] = students.count()
    ws_summary['A4'] = "Generated On:"
    ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # School-wise count
    ws_summary['A6'] = "School-wise Distribution:"
    ws_summary['A6'].font = Font(bold=True)
    row = 7
    
    schools = School.objects.all()
    for school in schools:
        school_students = students.filter(school=school).count()
        ws_summary[f'A{row}'] = school.name
        ws_summary[f'B{row}'] = school_students
        row += 1
    
    # Department-wise count
    ws_summary[f'A{row+1}'] = "Department-wise Distribution:"
    ws_summary[f'A{row+1}'].font = Font(bold=True)
    row += 2
    
    departments = Department.objects.all()
    for dept in departments:
        dept_students = students.filter(department=dept).count()
        ws_summary[f'A{row}'] = f"{dept.name} ({dept.school.name})"
        ws_summary[f'B{row}'] = dept_students
        row += 1
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Registered_Students_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response